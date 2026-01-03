from django.shortcuts import render,redirect
from django.http import request, HttpResponse
from .models import Product, Product_model
from .Forms.product_form import productForm, uploadFileForm
import pandas as pd 
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
import datetime
from datetime import datetime, date
from django.views.decorators.http import require_POST
from django.shortcuts import render, get_object_or_404
from .models import Product
from apps.complain_form.models import Event 
from django.db.models import  Sum
from apps.oem.models import ComplaintStatus, RepairCost, CustomerPricing

from django.core.paginator import Paginator
from django.db.models import OuterRef, Subquery, Sum, Value, DecimalField, CharField, DateTimeField
from django.db.models.functions import Coalesce
from django.db import IntegrityError
from openpyxl import Workbook
from openpyxl.utils import get_column_letter



@login_required(login_url='login')
# def create_product_view(request):
#     """Handle creation of a new Product via POST and redirect to product list.

#     On successful creation the user is redirected to the product list.
#     Validation errors are added to the messages framework.
#     """
#     if request.method == "POST":
#         form = productForm(request.POST)
#         if form.is_valid():
#             try:
#                 form.save()
#                 messages.success(request, "Product created successfully.")
#                 return redirect("product_list")
#             except IntegrityError:
#                 messages.error(request, "Tail Number already exists.")
#         else:
#             messages.error(request, "Tail Number already exists.")
#     else:
#         form = productForm()

#     return redirect('product_list')

def create_product_view(request):
    if request.method == "POST":
        form = productForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Product created successfully.")
        else:
            messages.error(request, "Please correct the errors in the form.")

    return redirect("product_list")



def product_list_view(request):
    """Display a filtered, paginated list of products.

    Supports filtering by model, sale status and warranty. Results are
    ordered newest-first so recently created products appear on top.
    Returns a Django `Page` object as `products` in the template context.
    """

    # Use a queryset so we can paginate efficiently and show newest first
    products_qs = Product.objects.all().order_by('-id')
    models = Product_model.objects.all()

    selected_model = request.GET.get("model")
    selected_status = request.GET.get("status")
    selected_warranty = request.GET.get("warranty")

    # Filter by product model
    if selected_model:
        products_qs = products_qs.filter(product_model_id=selected_model)

    # Filter by status
    if selected_status == "active":
        products_qs = products_qs.filter(active_status=True)
    elif selected_status == "obsolete":
        products_qs = products_qs.filter(active_status=False)

    # Warranty filtering using queryset filters (keeps it as a QuerySet for pagination)
    today = date.today()
    if selected_warranty == "active":
        products_qs = [p for p in products_qs if p.warranty_expiry_date >= today]
    elif selected_warranty == "expired":
        products_qs = [p for p in products_qs if p.warranty_expiry_date < today]
    # Pagination: 10 items per page (adjustable)
    page_number = request.GET.get('page', 1)
    paginator = Paginator(products_qs, 15)
    products_page = paginator.get_page(page_number)

    return render(request, "products/products_main_page.html", {
        "products": products_page,
        "models": models,
        "selected_model": selected_model,
        "selected_status": selected_status,
        "selected_warranty": selected_warranty,
    })


def edit_product_view(request, pk):
    product = get_object_or_404(Product, pk=pk)

    if request.method == "POST":
        model_id = request.POST.get('product_model')
        if model_id:
            product.product_model = get_object_or_404(Product_model, pk=model_id)

        product.contract_number = request.POST.get('contract_number')
        product.delivery_location = request.POST.get('delivery_location')
        product.army_command = request.POST.get('army_command')
        product.unit_name = request.POST.get('unit_name')
        product.formation = request.POST.get('formation')
        product.current_location = request.POST.get('current_location')
        product.remarks = request.POST.get('remarks')

        # ✅ ACTIVE STATUS FIX
        active_status = request.POST.get("active_status")
        product.active_status = True if active_status == "True" else False

        manufacture_date_str = request.POST.get('delivery_date')
        if manufacture_date_str:
            product.delivery_date = datetime.strptime(
                manufacture_date_str, '%Y-%m-%d'
            ).date()

        product.save()
        return redirect('product_list')

    return render(request, 'products/product_form.html', {
        'product': product,
        'product_models': Product_model.objects.all()
    })

import re
@login_required(login_url='login')
def import_products_view(request):

    if request.method != "POST":
        return redirect("product_list")

    if "import_file" not in request.FILES:
        messages.error(request, "⚠️ No file uploaded.")
        return redirect("product_list")

    file = request.FILES["import_file"]

    try:
        # Read file
        if file.name.endswith(".csv"):
            df = pd.read_csv(file, dtype=str)
        elif file.name.endswith((".xls", ".xlsx")):
            df = pd.read_excel(file, dtype=str)
        else:
            messages.error(request, "Invalid file type.")
            return redirect("product_list")

        # Normalize column names
        df.columns = df.columns.str.strip().str.lower()

        required_columns = [
            "tail_number",
            "model_name",
            "contract_number",
            "active_status",
            "delivery_date",
            "delivery_location",
            "warranty_period",
            "army_command",
            "unit_name",
            "formation",
        ]

        missing = [col for col in required_columns if col not in df.columns]
        if missing:
            messages.error(
                request,
                f"Missing required columns: {', '.join(missing)}"
            )
            return redirect("product_list")

        for index, row in df.iterrows():
            try:
                # ---------- Product Model ----------
                product_model = None
                model_name = str(row.get("model_name", "")).strip()
                if model_name:
                    product_model, _ = Product_model.objects.get_or_create(
                        model_name=model_name
                    )

                # ---------- Active Status ----------
                active_status = str(row.get("active_status", "")).strip().lower()
                active_status = active_status in ["true", "yes", "1"]

                # ---------- Delivery Date ----------
                delivery_date = None
                if pd.notna(row.get("delivery_date")):
                    delivery_date = pd.to_datetime(
                        row["delivery_date"], errors="coerce"
                    )
                    if delivery_date is not None:
                        delivery_date = delivery_date.date()

                # ---------- Warranty ----------
                warranty_period = 24
                warranty_raw = str(row.get("warranty_period", ""))
                numbers = re.findall(r"\d+", warranty_raw)
                if numbers:
                    warranty_period = int(numbers[0])

                # ---------- Save / Update ----------
                Product.objects.update_or_create(
                    tail_number=str(row["tail_number"]).strip(),
                    defaults={
                        "product_model": product_model,
                        "contract_number": str(row["contract_number"]).strip(),
                        "active_status": active_status,
                        "delivery_date": delivery_date,
                        "delivery_location": str(row["delivery_location"]).strip(),
                        "warranty_period": warranty_period,
                        "army_command": str(row["army_command"]).strip(),
                        "unit_name": str(row["unit_name"]).strip(),
                        "formation": str(row["formation"]).strip(),
                    }
                )

            except Exception as row_error:
                messages.error(
                    request,
                    f"Row {index + 1} failed: {row_error}"
                )
                continue

        messages.success(request, "✅ Products imported successfully.")
        return redirect("product_list")

    except Exception as e:
        messages.error(request, f"❌ Import failed: {e}")
        return redirect("product_list")



from dateutil.relativedelta import relativedelta 
@login_required(login_url='login')
def export_products_view(request):
    import pandas as pd
    from django.http import HttpResponse

    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")

    products = Product.objects.select_related("product_model").all()

    if from_date and to_date:
        products = products.filter(delivery_date__range=[from_date, to_date])

    # ✅ Build data explicitly (column → value)
    data = []
    for p in products:
        data.append({
            "Model Name": p.product_model.model_name if p.product_model else "",
            "Tail Number": p.tail_number,
            "Contract Number": p.contract_number,
            "Delivery Date": p.delivery_date,
            "Delivery Location": p.delivery_location,
            "Current Location": p.current_location,
            "Army Command": p.army_command,
            "Unit Name": p.unit_name,
            "Formation": p.formation,
            "Warranty (Months)": p.warranty_period,
            "Warranty Status": (
                "Expired" if p.delivery_date and
                (p.delivery_date + relativedelta(months=p.warranty_period)) < date.today()
                else "Active"
            ),
            "Remarks": p.remarks,
            "Obsolete": "Yes" if p.active_status else "No",
        })

    # ✅ Create DataFrame with clean columns
    df = pd.DataFrame(data)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="products.xlsx"'

    df.to_excel(response, index=False)

    return response


@require_POST
def delete_product_view(request, pk):
    product_obj = get_object_or_404(Product, id=pk)
    product_obj.delete()
    return redirect('product_list')



def repair_history_view(request, pk):
    product = Product.objects.get(pk=pk)
    page_number = request.GET.get('page', 1)
    tail_number = product.tail_number

    # Fetch all events for this tail number
    events_qs = Event.objects.filter(tail_number=tail_number)
    repairs_count = events_qs.count()
    first_event = events_qs.first()

    # ---- Subqueries ----
    latest_status_subquery = ComplaintStatus.objects.filter(
        event=OuterRef("pk")
    ).order_by("-updated_at").values("status")[:1]

    latest_date_subquery = ComplaintStatus.objects.filter(
        event=OuterRef("pk")
    ).order_by("-updated_at").values("updated_at")[:1]

    repair_cost_subquery = RepairCost.objects.filter(
        event=OuterRef("pk")
    ).values("event").annotate(total=Sum("repair_cost")).values("total")[:1]

    client_charge_subquery = CustomerPricing.objects.filter(
        event=OuterRef("pk")
    ).values("total_price")[:1]

    # ---- Annotate all in one queryset ----
    events = events_qs.annotate(
    last_status=Subquery(latest_status_subquery, output_field=CharField()),
    last_repair_date=Subquery(latest_date_subquery, output_field=DateTimeField()),
    total_repair_cost=Subquery(repair_cost_subquery, output_field=DecimalField()),
    client_charge=Subquery(client_charge_subquery, output_field=DecimalField()),
    ).order_by('-id')


    # ---- Totals for summary cards ----
    event_ids = events_qs.values_list("id", flat=True)

    total_repair_cost = (
        RepairCost.objects.filter(event_id__in=event_ids)
        .aggregate(total=Coalesce(Sum("repair_cost"), Value(0, output_field=DecimalField())))
        ["total"]
    )

    total_client_charge = (
        CustomerPricing.objects.filter(event_id__in=event_ids)
        .aggregate(total=Coalesce(Sum("total_price"), Value(0, output_field=DecimalField())))
        ["total"]
    )

    net_profit = total_client_charge - total_repair_cost
    avg_cost = total_repair_cost / repairs_count if repairs_count else 0
    margin = (net_profit / total_client_charge * 100) if total_client_charge > 0 else 0
    # Paginate EVENTS (repair history)
    paginator = Paginator(events, 10)
    event_page = paginator.get_page(page_number)

    context = {
    "product": product,
    "events": event_page,   # <-- paginated events
    "repairs_count": repairs_count,
    "first_event": first_event,
    "total_repair_cost": total_repair_cost,
    "total_client_charge": total_client_charge,
    "net_profit": net_profit,
    "avg_cost": avg_cost,
    "margin": margin,
}

    return render(request, "complaints/repair_history.html", context)

def repair_history_export_view(request, pk):
    # Get product & tail number
    product = get_object_or_404(Product, pk=pk)
    tail_number = product.tail_number

    # ---- Base queryset ----
    events_qs = Event.objects.filter(tail_number=tail_number)
    repairs_count = events_qs.count()

    # ---- Subqueries (same as in your repair_history_view) ----
    latest_status_subquery = ComplaintStatus.objects.filter(
        event=OuterRef("pk")
    ).order_by("-updated_at").values("status")[:1]

    latest_date_subquery = ComplaintStatus.objects.filter(
        event=OuterRef("pk")
    ).order_by("-updated_at").values("updated_at")[:1]

    repair_cost_subquery = RepairCost.objects.filter(
        event=OuterRef("pk")
    ).values("event").annotate(total=Sum("repair_cost")).values("total")[:1]

    client_charge_subquery = CustomerPricing.objects.filter(
        event=OuterRef("pk")
    ).values("total_price")[:1]

    # ---- Annotate events (no pagination here) ----
    events = events_qs.annotate(
        last_status=Subquery(latest_status_subquery, output_field=CharField()),
        last_repair_date=Subquery(latest_date_subquery, output_field=DateTimeField()),
        total_repair_cost=Subquery(repair_cost_subquery, output_field=DecimalField()),
        client_charge=Subquery(client_charge_subquery, output_field=DecimalField()),
    ).order_by("-id")

    # ---- Totals for summary ----
    event_ids = events_qs.values_list("id", flat=True)

    total_repair_cost = (
        RepairCost.objects.filter(event_id__in=event_ids)
        .aggregate(total=Coalesce(Sum("repair_cost"), Value(0, output_field=DecimalField())))
        ["total"]
    )

    total_client_charge = (
        CustomerPricing.objects.filter(event_id__in=event_ids)
        .aggregate(total=Coalesce(Sum("total_price"), Value(0, output_field=DecimalField())))
        ["total"]
    )

    net_profit = total_client_charge - total_repair_cost if total_client_charge is not None else 0
    avg_cost = total_repair_cost / repairs_count if repairs_count else 0
    margin = (net_profit / total_client_charge * 100) if total_client_charge and total_client_charge > 0 else 0

    # ==========================
    #   Create XLSX workbook
    # ==========================
    wb = Workbook()

    # ---------- Sheet 1: Repair History ----------
    ws = wb.active
    ws.title = "Repair History"

    headers = [
        "Event ID",
        "Unique Token",
        "Complaint Date",
        "Last Status",
        "Last Repair Date",
        "Total Repair Cost",
        "Client Charge",
    ]
    ws.append(headers)

    for event in events:
        # created_at may be timezone-aware → strip tz
        created = getattr(event, "date_of_occurrence", None)
        if created is not None and hasattr(created, "tzinfo") and created.tzinfo is not None:
            created = created.replace(tzinfo=None)

        # last_repair_date is from Subquery → also timezone-aware
        last_date = event.last_repair_date
        if last_date is not None and hasattr(last_date, "tzinfo") and last_date.tzinfo is not None:
            last_date = last_date.replace(tzinfo=None)

        ws.append([
            event.id,
            event.unique_token,  # assuming Event has unique_token
            created,
            event.last_status,
            last_date,
            event.total_repair_cost or 0,
            event.client_charge or 0,
        ])

    # Optional: auto-fit column width (simple version)
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            value = cell.value
            if value is None:
                continue
            value = str(value)
            if len(value) > max_length:
                max_length = len(value)
        ws.column_dimensions[col_letter].width = max_length + 2

    # ---------- Sheet 2: Summary ----------
    summary_ws = wb.create_sheet(title="Summary")
    summary_ws.append(["Product", str(product)])
    summary_ws.append(["Tail Number", tail_number])
    summary_ws.append(["Total Repairs", repairs_count])
    summary_ws.append(["Total Repair Cost", float(total_repair_cost or 0)])
    summary_ws.append(["Total Client Charge", float(total_client_charge or 0)])
    summary_ws.append(["Net Profit", float(net_profit or 0)])
    summary_ws.append(["Average Repair Cost", float(avg_cost or 0)])
    summary_ws.append(["Margin (%)", float(margin or 0)])

    # ==========================
    #   HTTP Response
    # ==========================
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"repair_history_{tail_number}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response



